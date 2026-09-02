"""Таблиця чатів і linked-груп із сирого дампа border_resolve_raw.json.

Один рядок = одне місце, де ЛЮДИ можуть писати:
  * кандидат, який сам є чатом/групою;
  * linked-група каналу (це і є «коментарі під постами»).
Канал без linked-групи в основний аркуш не потрапляє — коментарів там немає
взагалі; такі зібрані окремим аркушем як негативний результат.

Запуск (на хості, поза контейнером):  python3 backend/_dir/border_export_chats.py
"""
import json
import pathlib
import collections

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parents[2]
RAW = ROOT / "backend/_dir/border_resolve_raw.json"
PTS = json.loads((ROOT / "backend/analysis/fixtures/border_points.json").read_text())["points"]
OUT = ROOT / "docs/border-chats.xlsx"

FONT = "Arial"
HDR = PatternFill("solid", fgColor="1F4E79")
YES = PatternFill("solid", fgColor="C6EFCE")   # список учасників відкритий
NO = PatternFill("solid", fgColor="FFC7CE")
order = {p["point"]: i for i, p in enumerate(PTS)}


def yn(v):
    return "так" if v is True else ("ні" if v is False else "—")


def row_from(ent, full, kind, rec, parent=None):
    """Один рядок таблиці з сирих entity+ChannelFull."""
    u = ent.get("username") or ""
    return {
        "Регіон": rec["region"],
        "Точка": rec["point"] or "(регіон загалом)",
        "Що це": kind,
        "Назва": ent.get("title") or "",
        "Юзернейм": u,
        "Посилання": f"https://t.me/{u}" if u else "",
        "Учасників": full.get("participants_count"),
        "Список учасників відкритий": yn(full.get("can_view_participants")),
        "Slowmode, с": full.get("slowmode_seconds") or "",
        "Треба вступити щоб писати": yn(ent.get("join_to_send")),
        "Заявка на вступ": yn(ent.get("join_request")),
        "Мегагрупа": yn(ent.get("megagroup")),
        "Форум (теми)": yn(ent.get("forum")),
        "Обмежений (restricted)": yn(ent.get("restricted")),
        "Батьківський канал": parent or "",
        "Підписників каналу": (rec.get("full", {}).get("full_chat", {})
                               .get("participants_count") if parent else ""),
        "tg_id": ent.get("id"),
        "Опис": (full.get("about") or "").replace("\n", " ")[:300],
    }


raw = json.loads(RAW.read_text())
chats, no_comments, errors = [], [], []

for rec in raw.values():
    if not rec.get("ok"):
        errors.append({"Юзернейм": rec["username"], "Точка": rec["point"] or "",
                       "Регіон": rec["region"], "Помилка": rec.get("error", "")})
        continue
    ent = rec.get("entity") or {}
    full = (rec.get("full") or {}).get("full_chat") or {}

    if rec.get("linked_entity"):               # канал -> його група обговорення
        l_ent = rec["linked_entity"]
        l_full = (rec.get("linked_full") or {}).get("full_chat") or {}
        chats.append(row_from(l_ent, l_full, "група обговорення каналу", rec,
                              parent=ent.get("username") or ent.get("title") or ""))
    elif ent.get("megagroup") or ent.get("gigagroup"):   # сам по собі чат
        chats.append(row_from(ent, full, "самостійний чат", rec))
    else:
        # Розрізняємо два РІЗНІ випадки, які легко сплутати:
        #   коментарів немає взагалі  vs  коментарі є, але група приватна.
        # Другий виглядає як джерело, а насправді недоступний без вступу.
        has_linked = bool(full.get("linked_chat_id"))
        if rec.get("linked_error"):
            state = ("група обговорення ПРИВАТНА"
                     if "Private" in rec["linked_error"] else "linked-група не відкрилась")
        elif has_linked:
            state = "linked-група не зчиталась"
        else:
            state = "коментарів немає"
        no_comments.append({
            "Регіон": rec["region"], "Точка": rec["point"] or "(регіон загалом)",
            "Назва": ent.get("title") or "", "Юзернейм": ent.get("username") or "",
            "Підписників": full.get("participants_count"),
            "Стан": state,
            "linked_chat_id": full.get("linked_chat_id") or "",
            "Деталі": (rec.get("linked_error") or "")[:120],
        })

chats.sort(key=lambda r: (order.get(r["Точка"], 999), -(r["Учасників"] or 0)))

wb = Workbook()


def sheet(ws, rows, widths, fills=None):
    if not rows:
        ws.append(["(порожньо)"])
        return
    head = list(rows[0].keys())
    ws.append(head)
    for cell in ws[1]:
        cell.font = Font(name=FONT, bold=True, color="FFFFFF")
        cell.fill = HDR
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows:
        ws.append([r[h] for h in head])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name=FONT)
            cell.alignment = Alignment(vertical="top")
    if fills:
        col = head.index(fills) + 1
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            v = row[0].value
            row[0].fill = YES if v == "так" else (NO if v == "ні" else PatternFill())
    for i, h in enumerate(head, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 16)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


ws = wb.active
ws.title = "Чати і linked-групи"
sheet(ws, chats, {"Регіон": 24, "Точка": 17, "Що це": 24, "Назва": 40, "Юзернейм": 24,
                  "Посилання": 30, "Учасників": 11, "Список учасників відкритий": 15,
                  "Slowmode, с": 11, "Треба вступити щоб писати": 15,
                  "Заявка на вступ": 12, "Мегагрупа": 11, "Форум (теми)": 11,
                  "Обмежений (restricted)": 12, "Батьківський канал": 24,
                  "Підписників каналу": 13, "tg_id": 15, "Опис": 60},
      fills="Список учасників відкритий")
sheet(wb.create_sheet("Канали без доступних коментарів"), no_comments,
      {"Регіон": 24, "Точка": 17, "Назва": 40, "Юзернейм": 24, "Підписників": 12,
       "Стан": 28, "linked_chat_id": 15, "Деталі": 46})
sheet(wb.create_sheet("Помилки резолву"), errors,
      {"Юзернейм": 26, "Точка": 17, "Регіон": 24, "Помилка": 70})

wb.save(OUT)
opened = sum(1 for r in chats if r["Список учасників відкритий"] == "так")
print(f"чатів і linked-груп: {len(chats)}  (список учасників відкритий у {opened})")
import collections as _c
print("канали без доступних коментарів:",
      dict(_c.Counter(r["Стан"] for r in no_comments)))
print(f"усього таких: {len(no_comments)} | помилок резолву: {len(errors)}")
print(f"-> {OUT}")
